"""
商品コンセプトシート生成スクリプト (generate_concept_sheet.py)

Claude Code が競合ページをWebFetchで読み取り分析した結果を、
「商品コンセプトシート」テンプレート（template/商品コンセプトシート_フォーマット.xlsx）に
そのまま流し込んで .xlsx を生成する。

Groq/Claude APIやGoogle認証（service_account.json）は一切不要。
必要なのは openpyxl のみ（pip install openpyxl）。

【使い方】
    python generate_concept_sheet.py <入力JSONファイル> [出力先ディレクトリ]

【入力JSONの形式】
    {
      "title": "タイトル（仮）",
      "one_line": "この商品をひとことで言うと",
      "target": "ターゲット（年代/性別など）",
      "problems": "ターゲットの悩み/困りごと",
      "emotional_job": "感情ジョブ",
      "social_job": "社会ジョブ",
      "market_research": "市場調査（市場規模・訴求トレンド・空白地帯など）",
      "competitors": [
        {
          "name": "商品名／ブランド／販売元", "url": "URL",
          "price": "価格", "volume": "容量", "sales": "主要な売り場や売上",
          "features": "商品特長", "appeal": "メイン訴求",
          "reasons": "選ばれる理由", "complaints": "不満な声"
        }
        # 最大5件（①〜⑤）
      ],
      "entry_detail": "参入余地（勝ち筋）",
      "main_appeal": "メイン訴求（自社）",
      "product_strengths": "商品の強み（この商品でなければならない理由）",
      "product_form": "商品形状",
      "price_volume": "売価/容量",
      "sales_channel": "想定売り場",
      "key_ingredients": "キー成分",
      "bottleneck": "ボトルネックとそれを超える策",
      "cancellation_reason": "解雇（解約）条件",
      "author": "作成者（省略可、既定値: Claude Code）",
      "version": "Ver.（省略可、既定値: 01）"
    }
"""

import sys
import json
import os
from datetime import date

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "商品コンセプトシート_フォーマット.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

# 競合①〜⑤の開始行（各6行で1ブロック）
COMPETITOR_ROWS = [34, 40, 46, 52, 58]


def fill_sheet(ws, data: dict) -> None:
    ws["D3"] = date.today().strftime("%Y/%m/%d")
    ws["M3"] = data.get("author", "Claude Code")
    ws["AJ3"] = data.get("version", "01")

    ws["J6"] = data.get("title", "")
    ws["J7"] = data.get("one_line", "")
    ws["J8"] = data.get("target", "")
    ws["J9"] = data.get("problems", "")
    ws["J15"] = data.get("emotional_job", "")
    ws["J18"] = data.get("social_job", "")

    ws["A23"] = data.get("market_research", "")

    for i, comp in enumerate(data.get("competitors", [])[:5]):
        r = COMPETITOR_ROWS[i]
        ws[f"I{r}"] = comp.get("name", "")
        ws[f"AF{r}"] = comp.get("url", "")
        ws[f"D{r + 1}"] = comp.get("price", "")
        ws[f"J{r + 1}"] = comp.get("volume", "")
        ws[f"S{r + 1}"] = comp.get("sales", "")
        ws[f"E{r + 2}"] = comp.get("features", "")
        ws[f"E{r + 3}"] = comp.get("appeal", "")
        ws[f"F{r + 4}"] = comp.get("reasons", "")
        ws[f"E{r + 5}"] = comp.get("complaints", "")

    ws["E66"] = data.get("entry_detail", "")
    ws["E73"] = data.get("main_appeal", "")
    ws["E74"] = data.get("product_strengths", "")
    ws["E81"] = data.get("product_form", "")
    ws["Z81"] = data.get("price_volume", "")
    ws["E82"] = data.get("sales_channel", "")
    ws["R82"] = data.get("key_ingredients", "")
    ws["J83"] = data.get("bottleneck", "")
    ws["J86"] = data.get("cancellation_reason", "")


def main() -> None:
    if len(sys.argv) < 2:
        print("使い方: python generate_concept_sheet.py <入力JSONファイル> [出力先ディレクトリ]")
        sys.exit(1)

    json_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else OUTPUT_DIR
    os.makedirs(output_dir, exist_ok=True)

    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["商品コンセプトシート"]
    fill_sheet(ws, data)

    title = (data.get("title") or "商品企画").strip()
    safe_title = "".join(c for c in title if c not in '\\/:*?"<>|')[:20]
    filename = f"【{date.today().strftime('%m%d')}】{safe_title}.xlsx"
    output_path = os.path.join(output_dir, filename)
    wb.save(output_path)
    print(f"生成完了: {output_path}")


if __name__ == "__main__":
    main()
